import os
import openpyxl

#获取数据

path = r"C:\Users\sudoWF\Desktop"   # 指定文件路径（以桌面为例）
os.chdir(path)  # 修改工作路径

workbook = openpyxl.load_workbook('test2.xlsx')  # 打开原始Excel    # 返回一个workbook数据类型的值
#print(workbook.sheetnames)  # 打印Excel表中的所有表

sheet = workbook['Sheet1']  # 获取指定sheet表
#sheet = workbook.active     # 获取活动表
print(sheet)
print(sheet.dimensions)     # 获取表格的尺寸大小

#------------------------------------------

# 按行(rows)获取表单中的所有单元格，每一行的数据放到一个元组中
res = list(sheet.rows)

# 获取excel表格中的第一行的数据，作为字典的key，生成一个list列表
title = [i.value for i in res[0]]

# 作为每个字典的容器
cases = []
# 遍历第一行以外的所有行
for item in res[1:]:
    # 获取每行的数据
    dataline = [i.value for i in item]
    # 把遍历的每行数据与第一行title数据打包成字典
    dicline = dict(zip(title, dataline))
    cases.append(dicline)

# print(cases)

# print(cases[0]['姓名'])

new_cases = []
a=0
new_cases.append(cases[a])
a += 1

# dz = cases[a]['住房单元地址']
# dz0 = cases[a-1]['住房单元地址']
# print(dz)
# print(dz0)
# print(dz==dz0)

# 找出每一户在表格中排在第一位的人(户主)
while a < 16:   #这里是Excel表的行数-1
    dz1 = cases[a]['住房单元地址']
    dz0 = cases[a-1]['住房单元地址']
    if dz1 == dz0:
        a+=1
    else:
        new_cases.append(cases[a])
        a+=1

print(new_cases)


#------------------------------------------

#保存为新文件

from openpyxl import Workbook
def inputexcel(inputdata,outputdata):
    wb = Workbook()
    sheet = wb.active
    fd = inputdata[0]
    for zm,i in list(zip([chr(letter).upper() for letter in range(65, 91)],range(len(list(fd.keys()))))):
        sheet[zm+str(1)].value = list(fd.keys())[i]
    j = 2
    for item in inputdata:
        for zm, key in list(zip([chr(letter).upper() for letter in range(65, 91)], list(fd.keys()))):
            sheet[zm+str(j)] = item[key]
        j += 1
    wb.save(outputdata)

inputexcel(new_cases,'./output.xlsx')

#wb.save('output.xlsx')

